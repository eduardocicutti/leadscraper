import io
from contextlib import asynccontextmanager
import os
import logging
import multiprocessing
import traceback
import re
import time
import threading
import urllib.parse
from pathlib import Path
from datetime import datetime
from fastapi import FastAPI
from fastapi.middleware.cors import CORSMiddleware
from fastapi.responses import HTMLResponse, StreamingResponse, JSONResponse
from pydantic import BaseModel
from sqlmodel import Field, SQLModel, create_engine, Session, Relationship, select

LOG_DIR = Path(os.getenv("DB_PATH", "banco.db")).resolve().parent
LOG_DIR.mkdir(parents=True, exist_ok=True)
logging.basicConfig(
    level=logging.INFO,
    format="%(asctime)s [%(levelname)s] %(name)s: %(message)s",
    handlers=[
        logging.StreamHandler(),
        logging.FileHandler(LOG_DIR / "backend.log", encoding="utf-8"),
    ],
)
logger = logging.getLogger("lead_scraper")

class ScrapeRequest(BaseModel):
    segmento: str
    cidade: str
    estado: str
    max_results: int = 30
    prospectador: str = ""
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.worksheet.datavalidation import DataValidation

@asynccontextmanager
async def lifespan(app: FastAPI):
    logger.info("FastAPI lifespan startup triggered")
    init_db(os.getenv("DB_PATH", "banco.db"))
    try:
        yield
    finally:
        logger.info("FastAPI lifespan shutdown triggered")

app = FastAPI(lifespan=lifespan)
app.add_middleware(
    CORSMiddleware,
    allow_origins=[
        "http://localhost:1420",
        "http://127.0.0.1:1420",
        "tauri://localhost",
        "http://tauri.localhost",
    ],
    allow_methods=["*"],
    allow_headers=["*"],
)
jobs: dict = {}
engine = None
db_lock = threading.Lock()


class SearchHistory(SQLModel, table=True):
    id: int | None = Field(default=None, primary_key=True)
    keyword: str
    city: str
    state: str
    status: str
    leads_found: int = 0
    prospectador: str
    created_at: datetime = Field(default_factory=datetime.now)

    leads: list["Lead"] = Relationship(
        back_populates="search_history",
        sa_relationship_kwargs={"cascade": "all, delete-orphan"},
    )


class Lead(SQLModel, table=True):
    id: int | None = Field(default=None, primary_key=True)
    search_history_id: int = Field(foreign_key="searchhistory.id")
    company_name: str
    phone: str | None = None
    is_whatsapp: bool = False
    whatsapp_link: str | None = None
    website: str | None = None
    address: str | None = None
    rating: float | None = None
    review_count: int | None = None
    category: str | None = None
    porte: str | None = None
    score: int = 0
    classificacao: str | None = None
    maps_url: str | None = None

    search_history: SearchHistory = Relationship(back_populates="leads")


def init_db(db_path: str) -> None:
    global engine
    try:
        resolved_path = Path(db_path).expanduser().resolve()
        resolved_path.parent.mkdir(parents=True, exist_ok=True)
        logger.info("Opening database at %s", resolved_path)
        engine = create_engine(
            f"sqlite:///{resolved_path}",
            connect_args={"check_same_thread": False},
        )
        SQLModel.metadata.create_all(engine)
        logger.info("Database initialized at %s", resolved_path)
    except Exception:
        logger.exception("Failed to initialize database at %s", db_path)
        traceback.print_exc()
        raise



def ensure_db_ready() -> None:
    if engine is None:
        init_db(os.getenv("DB_PATH", "banco.db"))


def save_search_history(
    job_id: str,
    segmento: str,
    cidade: str,
    estado: str,
    prospectador: str,
) -> int:
    try:
        ensure_db_ready()
        record = SearchHistory(
            keyword=segmento,
            city=cidade,
            state=estado,
            status="running",
            prospectador=prospectador,
        )
        with db_lock:
            with Session(engine) as session:
                session.add(record)
                session.commit()
                session.refresh(record)
                logger.info("Search history created id=%s job_id=%s", record.id, job_id)
                return record.id
    except Exception:
        logger.exception("Failed to save search history for job %s", job_id)
        traceback.print_exc()
        raise


def save_leads_batch(history_id: int, leads: list[dict]) -> None:
    try:
        ensure_db_ready()
        with db_lock:
            with Session(engine) as session:
                history = session.get(SearchHistory, history_id)
                if history is None:
                    logger.warning("Search history id=%s not found while saving leads", history_id)
                    return

                for lead in leads:
                    session.add(
                        Lead(
                            search_history_id=history_id,
                            company_name=lead.get("nome", ""),
                            phone=lead.get("telefone"),
                            is_whatsapp=lead.get("is_whatsapp", False),
                            whatsapp_link=lead.get("whatsapp_link") or None,
                            website=lead.get("site"),
                            address=lead.get("endereco"),
                            rating=lead.get("nota"),
                            review_count=lead.get("avaliacoes"),
                            category=lead.get("categoria"),
                            porte=lead.get("porte"),
                            score=lead.get("score", 0),
                            classificacao=lead.get("classificacao"),
                            maps_url=lead.get("url_maps"),
                        )
                    )

                history.leads_found = len(leads)
                history.status = "done"
                session.add(history)
                session.commit()
                logger.info("Saved %s leads for history id=%s", len(leads), history_id)
    except Exception:
        logger.exception("Failed to save leads for history id=%s", history_id)
        traceback.print_exc()
        raise


def update_search_history_status(history_id: int, status: str, leads_found: int = 0) -> None:
    try:
        ensure_db_ready()
        with db_lock:
            with Session(engine) as session:
                history = session.get(SearchHistory, history_id)
                if history is None:
                    logger.warning("Search history id=%s not found while updating status", history_id)
                    return
                history.status = status
                history.leads_found = leads_found
                session.add(history)
                session.commit()
                logger.info("Updated history id=%s status=%s leads_found=%s", history_id, status, leads_found)
    except Exception:
        logger.exception("Failed to update history id=%s status=%s", history_id, status)
        traceback.print_exc()
        raise
# Segmentos com alto potencial digital
HIGH_FIT_SEGMENTS = [
    "clínica", "consultório", "odontologia", "dentista", "médico", "saúde",
    "advocacia", "advogado", "escritório", "contabilidade", "contador",
    "imobiliária", "imóveis", "corretor",
    "academia", "personal", "fitness",
    "restaurante", "lanchonete", "delivery", "alimentação",
    "hotel", "pousada", "turismo",
    "escola", "curso", "educação", "colégio",
    "loja", "comércio", "varejo", "e-commerce",
    "construtora", "engenharia", "arquitetura",
    "pet", "veterinário",
    "beleza", "estética", "salão", "barbearia",
    "tecnologia", "software", "startup",
    "indústria", "manufatura", "fábrica",
    "logística", "transportadora",
    "financeiro", "investimento", "seguro",
]

PORTE_KEYWORDS = {
    "MEI / Autônomo": ["autônomo", "individual", "freelancer", "mei"],
    "Micro empresa": ["micro", "microempresa"],
    "Pequena empresa": ["pequena", "ltda", "me "],
    "Média empresa": ["média", "médio porte"],
    "Grande empresa": ["grande", "s/a", "sa ", "grupo ", "holding", "nacional", "rede "],
}

def classify_porte(nome: str, categoria: str, avaliacoes: int) -> str:
    texto = f"{nome} {categoria}".lower()
    for porte, kws in PORTE_KEYWORDS.items():
        if any(kw in texto for kw in kws):
            return porte
    # Heurística por avaliações
    if not avaliacoes:
        return "Micro empresa"
    if avaliacoes >= 500: return "Grande empresa"
    if avaliacoes >= 200: return "Média empresa"
    if avaliacoes >= 50:  return "Pequena empresa"
    return "Micro empresa"

def score_lead(lead: dict) -> tuple:
    """
    Sistema de score comercial — máx 100 pontos
    
    Dimensão 1 — Presença digital atual (0-30 pts)
      Sem site: +25 (dor latente → oportunidade de site/lp)
      Site ruim / apenas redes sociais: +15
      Tem site: +5 (pode querer sistema/app/redesign)
    
    Dimensão 2 — Reputação e volume (0-20 pts)
      Avaliações ≥ 100: +20 (empresa ativa, investe em marketing)
      Avaliações 50-99: +15
      Avaliações 10-49: +10
      Avaliações 1-9: +5
    
    Dimensão 3 — Nota Google (0-15 pts)
      Nota ≥ 4.5: +15 (empresa de qualidade, vale prospectar)
      Nota 4.0-4.4: +10
      Nota 3.5-3.9: +5
    
    Dimensão 4 — Fit de segmento (0-20 pts)
      Segmento HIGH_FIT: +20
      Tem telefone/WhatsApp: +10 (contatável)
    
    Dimensão 5 — Porte (0-15 pts)
      Grande empresa: +15
      Média empresa: +12
      Pequena empresa: +10
      Micro empresa: +5
      MEI: +2
    
    Temperatura:
      🔥 Quente  → score ≥ 70
      🟡 Morno   → score 45-69
      ❄️ Frio    → score < 45
    """
    score = 0
    avaliacoes = lead.get("avaliacoes") or 0
    nota       = lead.get("nota") or 0.0
    tem_site   = bool(lead.get("site"))
    tem_tel    = bool(lead.get("telefone"))
    categoria  = (lead.get("categoria") or "").lower()
    nome       = (lead.get("nome") or "").lower()
    porte      = lead.get("porte") or ""

    # Dim 1 — Presença digital
    if not tem_site:
        score += 25
    else:
        score += 5

    # Dim 2 — Avaliações
    if avaliacoes >= 100:   score += 20
    elif avaliacoes >= 50:  score += 15
    elif avaliacoes >= 10:  score += 10
    elif avaliacoes >= 1:   score += 5

    # Dim 3 — Nota
    if nota >= 4.5:   score += 15
    elif nota >= 4.0: score += 10
    elif nota >= 3.5: score += 5

    # Dim 4 — Fit de segmento
    if any(kw in categoria or kw in nome for kw in HIGH_FIT_SEGMENTS):
        score += 20
    if tem_tel:
        score += 10

    # Dim 5 — Porte
    porte_pts = {"Grande empresa": 15, "Média empresa": 12, "Pequena empresa": 10,
                 "Micro empresa": 5, "MEI / Autônomo": 2}
    score += porte_pts.get(porte, 5)

    score = min(score, 100)

    if score >= 70:   clf = "🔥 Quente"
    elif score >= 45: clf = "🟡 Morno"
    else:             clf = "❄️ Frio"

    return clf, score

def build_whatsapp_link(telefone: str, nome_empresa: str) -> str:
    if not telefone:
        return ""
    digits = re.sub(r'\D', '', telefone)
    if not digits.startswith("55"):
        if len(digits) == 10 or len(digits) == 11:
            digits = "55" + digits
        else:
            digits = "55" + digits
    msg = (
        f"Olá! Tudo bem? 😊\n\n"
        f"Entrei em contato porque identificamos que a *{nome_empresa}* "
        f"pode se beneficiar com soluções digitais personalizadas — seja um site institucional, "
        f"landing page, sistema web ou aplicativo mobile.\n\n"
        f"Posso te apresentar como podemos ajudar? Seria rapidinho! 🚀"
    )
    encoded = urllib.parse.quote(msg)
    return f"https://wa.me/{digits}?text={encoded}"

def scrape_worker(job_id: str, segmento: str, cidade: str, estado: str,
                  max_results: int, prospectador: str):
    from selenium import webdriver
    from selenium.webdriver.edge.service import Service
    from selenium.webdriver.edge.options import Options
    from selenium.webdriver.common.by import By
    from webdriver_manager.microsoft import EdgeChromiumDriverManager

    logger.info("Starting scraper job_id=%s segmento=%s cidade=%s estado=%s max_results=%s", job_id, segmento, cidade, estado, max_results)
    history_id: int | None = None
    driver = None
    leads: list[dict] = []

    def set_error(message: str) -> None:
        jobs[job_id]["status"] = "error"
        jobs[job_id]["log"] = message
        logger.error("Job %s failed: %s", job_id, message)

    def find_text(default: str | None, description: str, callback):
        try:
            return callback()
        except Exception as exc:
            logger.debug("Could not extract %s for job %s: %s", description, job_id, exc, exc_info=True)
            return default

    try:
        jobs[job_id]["status"] = "running"
        jobs[job_id]["log"] = "Iniciando Edge..."
        history_id = save_search_history(job_id, segmento, cidade, estado, prospectador)

        options = Options()
        options.add_argument("--headless=new")
        options.add_argument("--no-sandbox")
        options.add_argument("--disable-dev-shm-usage")
        options.add_argument("--disable-gpu")
        options.add_argument("--window-size=1280,900")
        options.add_argument("--lang=pt-BR")
        options.add_argument("--disable-blink-features=AutomationControlled")
        options.add_experimental_option("excludeSwitches", ["enable-automation"])
        options.add_experimental_option("useAutomationExtension", False)
        options.add_argument("user-agent=Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/124.0.0.0 Safari/537.36")

        try:
            logger.info("Initializing Edge webdriver for job %s", job_id)
            driver = webdriver.Edge(
                service=Service(EdgeChromiumDriverManager().install()),
                options=options,
            )
            driver.execute_script("Object.defineProperty(navigator, 'webdriver', {get: () => undefined})")
        except Exception as exc:
            logger.exception("Failed to initialize Edge webdriver for job %s", job_id)
            traceback.print_exc()
            set_error(f"Erro ao iniciar Edge: {str(exc)[:180]}")
            if history_id is not None:
                update_search_history_status(history_id, "error")
            return

        query = f"{segmento} em {cidade} {estado}"
        url = f"https://www.google.com/maps/search/{query.replace(' ', '+')}"
        jobs[job_id]["log"] = f"Abrindo Google Maps: {query}"
        logger.info("Opening Google Maps for job %s: %s", job_id, query)
        driver.get(url)
        time.sleep(4)

        jobs[job_id]["log"] = "Carregando resultados..."
        for attempt in range(6):
            try:
                panel = driver.find_element(By.CSS_SELECTOR, '[role="feed"]')
                driver.execute_script("arguments[0].scrollBy(0, 1000)", panel)
                time.sleep(1.5)
            except Exception as exc:
                logger.warning("Could not scroll results panel on attempt %s for job %s: %s", attempt + 1, job_id, exc)
                break

        cards = driver.find_elements(By.CSS_SELECTOR, 'a[href*="/maps/place/"]')
        hrefs: list[str] = []
        seen: set[str] = set()
        for card in cards:
            try:
                href = card.get_attribute("href")
            except Exception as exc:
                logger.debug("Could not read maps card href for job %s: %s", job_id, exc, exc_info=True)
                continue
            if href and href not in seen and "/maps/place/" in href:
                seen.add(href)
                hrefs.append(href)
            if len(hrefs) >= max_results:
                break

        jobs[job_id]["total"] = len(hrefs)
        jobs[job_id]["log"] = f"Encontrou {len(hrefs)} estabelecimentos. Extraindo detalhes..."
        logger.info("Job %s found %s candidate URLs", job_id, len(hrefs))

        for i, href in enumerate(hrefs):
            try:
                driver.get(href)
                time.sleep(2.5)

                nome = find_text(
                    driver.title.replace(" - Google Maps", "").strip(),
                    "company name",
                    lambda: driver.find_element(By.CSS_SELECTOR, "h1").text.strip(),
                )

                nota = find_text(
                    None,
                    "rating",
                    lambda: float(driver.find_element(By.CSS_SELECTOR, 'div[jsaction*="pane.rating"] span[aria-hidden="true"]').text.replace(",", ".")),
                )
                if nota is None:
                    def rating_from_aria():
                        el = driver.find_element(By.CSS_SELECTOR, 'span[aria-label*="estrela"]')
                        match = re.search(r'(\d[.,]\d)', el.get_attribute("aria-label") or "")
                        return float(match.group(1).replace(",", ".")) if match else None
                    nota = find_text(None, "rating aria-label", rating_from_aria)

                def reviews_from_button():
                    el = driver.find_element(By.CSS_SELECTOR, 'button[jsaction*="pane.rating.moreReviews"] span')
                    nums = re.findall(r'\d+', el.text.replace(".", "").replace(",", ""))
                    return int(nums[0]) if nums else None
                avaliacoes = find_text(None, "review count", reviews_from_button)
                if avaliacoes is None:
                    def reviews_from_text():
                        el = driver.find_element(By.XPATH, '//span[contains(text(),"avaliações") or contains(text(),"avaliação")]')
                        nums = re.findall(r'\d+', el.text.replace(".", "").replace(",", ""))
                        return int(nums[0]) if nums else None
                    avaliacoes = find_text(None, "review count text", reviews_from_text)

                endereco = find_text(
                    None,
                    "address",
                    lambda: driver.find_element(By.CSS_SELECTOR, 'button[data-item-id="address"]').get_attribute("aria-label").replace("Endereço: ", ""),
                )
                if endereco is None:
                    endereco = find_text(
                        None,
                        "address tooltip",
                        lambda: driver.find_element(By.XPATH, '//button[@data-tooltip="Copiar endereço"]').get_attribute("aria-label"),
                    )

                telefone = find_text(
                    None,
                    "phone",
                    lambda: driver.find_element(By.CSS_SELECTOR, 'button[data-item-id*="phone"]').get_attribute("aria-label").replace("Número de telefone: ", "").strip(),
                )
                if telefone is None:
                    telefone = find_text(
                        None,
                        "phone aria-label",
                        lambda: driver.find_element(By.XPATH, '//button[contains(@aria-label,"telefone") or contains(@aria-label,"phone")]').get_attribute("aria-label").strip(),
                    )

                is_whatsapp = False
                if telefone:
                    digits = re.sub(r'\D', '', telefone)
                    if len(digits) >= 10:
                        local = digits[-9:] if len(digits) >= 9 else digits
                        is_whatsapp = local.startswith('9')

                site = find_text(
                    None,
                    "website",
                    lambda: driver.find_element(By.CSS_SELECTOR, 'a[data-item-id="authority"]').get_attribute("href"),
                )
                if site is None:
                    site = find_text(
                        None,
                        "website aria-label",
                        lambda: driver.find_element(By.XPATH, '//a[contains(@aria-label,"Site")]').get_attribute("href"),
                    )

                categoria = find_text(
                    None,
                    "category",
                    lambda: driver.find_element(By.CSS_SELECTOR, 'button.DkEaL').text.strip(),
                )

                porte = classify_porte(nome, categoria or segmento, avaliacoes or 0)
                whatsapp_link = build_whatsapp_link(telefone, nome) if is_whatsapp else ""

                lead = {
                    "nome": nome,
                    "categoria": categoria or segmento,
                    "nota": nota,
                    "avaliacoes": avaliacoes,
                    "endereco": endereco,
                    "telefone": telefone,
                    "is_whatsapp": is_whatsapp,
                    "whatsapp_link": whatsapp_link,
                    "site": site,
                    "url_maps": href,
                    "porte": porte,
                    "prospectador": prospectador,
                    "cidade": cidade,
                    "estado": estado,
                }
                clf, score = score_lead(lead)
                lead["classificacao"] = clf
                lead["score"] = score

                leads.append(lead)
                jobs[job_id]["leads"] = sorted(leads, key=lambda x: x["score"], reverse=True)
                jobs[job_id]["progress"] = i + 1
                jobs[job_id]["log"] = f"Extraindo {i+1}/{len(hrefs)}: {nome}"

            except Exception as exc:
                logger.exception("Failed to extract lead %s/%s for job %s", i + 1, len(hrefs), job_id)
                traceback.print_exc()
                jobs[job_id]["log"] = f"Pulando lead {i+1}: {str(exc)[:80]}"
                continue

        if leads:
            jobs[job_id]["leads"] = sorted(leads, key=lambda x: x["score"], reverse=True)
            jobs[job_id]["status"] = "done"
            jobs[job_id]["log"] = f"Concluído! {len(leads)} leads coletados."
            if history_id is not None:
                save_leads_batch(history_id, jobs[job_id]["leads"])
            logger.info("Job %s completed with %s leads", job_id, len(leads))
        else:
            set_error("Nenhum lead encontrado. Tente outro segmento ou cidade.")
            if history_id is not None:
                update_search_history_status(history_id, "error", 0)

    except Exception as exc:
        logger.exception("Unexpected scraper failure for job %s", job_id)
        traceback.print_exc()
        set_error(f"Erro: {str(exc)[:180]}")
        if history_id is not None:
            update_search_history_status(history_id, "error", len(jobs[job_id].get("leads", leads)))
    finally:
        if driver is not None:
            try:
                driver.quit()
            except Exception:
                logger.exception("Failed to close Edge webdriver for job %s", job_id)
                traceback.print_exc()

def generate_xlsx(leads: list, segmento: str, cidade: str, estado: str, prospectador: str) -> bytes:
    """
    Gera planilha no formato EXATO do modelo Julho.xlsx
    Colunas originais mantidas + novas colunas extras
    """
    wb = Workbook()
    ws = wb.active
    ws.title = "Página1"

    # Cores do modelo
    HEADER_FILL = PatternFill("solid", start_color="1C4587")  # Azul escuro do modelo
    WHITE_FONT  = Font(bold=True, color="FFFFFF", name="Calibri", size=11)
    thin = Side(style="thin", color="CCCCCC")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    # Cabeçalhos — mesma ordem do modelo + novas colunas
    headers = [
        "Responsável",        # A
        "Estágio atual",      # B
        "Nome da empresa",    # C
        "Telefone/WhatsApp",  # D
        "Responsável",        # E (contato na empresa)
        "Onde encontrou?",    # F
        "Resp",               # G (checkbox)
        "F1",                 # H (checkbox)
        "F2",                 # I (checkbox)
        "F3",                 # J (checkbox)
        "Ramo de atividade",  # K
        "Porte",              # L (novo)
        "Nota Google",        # M (novo)
        "Avaliações",         # N (novo)
        "Score",              # O (novo)
        "Temperatura",        # P (novo)
        "Link WhatsApp",      # Q (novo)
        "Endereço",           # R (novo)
        "Site",               # S (novo)
        "Google Maps",        # T (novo)
        "Cidade",             # U (novo)
    ]

    # Larguras das colunas (seguindo modelo)
    col_widths = {
        'A': 15.13, 'B': 17.0, 'C': 28.75, 'D': 17.0, 'E': 17.63,
        'F': 17.0,  'G': 7.63, 'H': 7.0,   'I': 7.0,  'J': 7.0,
        'K': 23.75, 'L': 18.0, 'M': 12.0,  'N': 12.0, 'O': 13.0,
        'P': 13.0,  'Q': 35.0, 'R': 40.0,  'S': 30.0, 'T': 40.0,
        'U': 15.0,
    }
    for col_letter, width in col_widths.items():
        ws.column_dimensions[col_letter].width = width

    # Linha de cabeçalho
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.font = WHITE_FONT
        cell.fill = HEADER_FILL
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = border
    ws.row_dimensions[1].height = 20

    # Data validations — igual ao modelo
    dv_responsavel = DataValidation(type="list",
        formula1='"Eduardo,Murilo,Sofia,Ydian,Responsável,Gabriel"', showDropDown=False)
    dv_estagio = DataValidation(type="list",
        formula1='"Qualificação,Mensagens,Briefing,Perdido"', showDropDown=False)
    dv_origem = DataValidation(type="list",
        formula1='"Google maps,Indicação,Instagram,Linkedin,Outros,B2B"', showDropDown=False)

    ws.add_data_validation(dv_responsavel)
    ws.add_data_validation(dv_estagio)
    ws.add_data_validation(dv_origem)

    # Cores por temperatura
    QUENTE_FILL = PatternFill("solid", start_color="FFE0E0")
    MORNO_FILL  = PatternFill("solid", start_color="FFF9C4")
    FRIO_FILL   = PatternFill("solid", start_color="E3F2FD")
    ALT_FACTOR  = "F8F8F8"  # linha alternada mais clara

    max_row = 1002
    for i, lead in enumerate(leads):
        row = i + 2
        clf = lead.get("classificacao", "")
        if "Quente" in clf:   row_fill = QUENTE_FILL
        elif "Morno" in clf:  row_fill = MORNO_FILL
        else:                  row_fill = FRIO_FILL

        tel = lead.get("telefone") or ""
        is_wa = lead.get("is_whatsapp", False)
        tel_display = f"{'📱 ' if is_wa else ''}{tel}"

        row_values = [
            prospectador,                    # A Responsável
            "Qualificação",                  # B Estágio inicial
            lead.get("nome", ""),            # C Nome da empresa
            tel_display,                     # D Telefone/WhatsApp
            "",                              # E Responsável (contato empresa)
            "Google maps",                   # F Onde encontrou
            False,                           # G Resp checkbox
            False,                           # H F1 checkbox
            False,                           # I F2 checkbox
            False,                           # J F3 checkbox
            lead.get("categoria", ""),       # K Ramo de atividade
            lead.get("porte", ""),           # L Porte
            lead.get("nota"),                # M Nota Google
            lead.get("avaliacoes"),          # N Avaliações
            lead.get("score", 0),            # O Score
            clf,                             # P Temperatura
            lead.get("whatsapp_link", ""),   # Q Link WhatsApp
            lead.get("endereco", ""),        # R Endereço
            lead.get("site", ""),            # S Site
            lead.get("url_maps", ""),        # T Google Maps
            f"{lead.get('cidade','')}/{lead.get('estado','')}", # U Cidade
        ]

        for col, val in enumerate(row_values, 1):
            cell = ws.cell(row=row, column=col, value=val)
            cell.fill = row_fill
            cell.border = border
            cell.alignment = Alignment(vertical="center", wrap_text=False)
            if col in (7, 8, 9, 10):  # checkboxes
                cell.alignment = Alignment(horizontal="center", vertical="center")
            if col == 3:  # nome empresa em negrito
                cell.font = Font(bold=True, name="Calibri", size=10)
            else:
                cell.font = Font(name="Calibri", size=10)

        ws.row_dimensions[row].height = 18

        # Aplicar data validations nas colunas A, B, F
        dv_responsavel.sqref = f"A2:A{max_row}"
        dv_estagio.sqref     = f"B2:B{max_row}"
        dv_origem.sqref      = f"F2:F{max_row}"

    # Freeze header
    ws.freeze_panes = "A2"

    # Aba de resumo / score legend
    ws2 = wb.create_sheet("Score")
    ws2["A1"] = "SISTEMA DE SCORE"
    ws2["A1"].font = Font(bold=True, size=14, color="1C4587")
    ws2.merge_cells("A1:C1")

    ws2["A3"] = "Dimensão"
    ws2["B3"] = "Critério"
    ws2["C3"] = "Pontos"
    for col in range(1, 4):
        ws2.cell(3, col).font = Font(bold=True, color="FFFFFF")
        ws2.cell(3, col).fill = HEADER_FILL
        ws2.cell(3, col).alignment = Alignment(horizontal="center")

    score_table = [
        ("Presença Digital", "Sem site (oportunidade clara de presença digital)", 25),
        ("Presença Digital", "Com site (pode querer sistema/app/redesign)", 5),
        ("Avaliações Google", "≥ 100 avaliações", 20),
        ("Avaliações Google", "50 – 99 avaliações", 15),
        ("Avaliações Google", "10 – 49 avaliações", 10),
        ("Avaliações Google", "1 – 9 avaliações", 5),
        ("Nota Google", "≥ 4.5 ⭐ (empresa de qualidade)", 15),
        ("Nota Google", "4.0 – 4.4 ⭐", 10),
        ("Nota Google", "3.5 – 3.9 ⭐", 5),
        ("Fit de Segmento", "Segmento com alto potencial digital", 20),
        ("Fit de Segmento", "Tem telefone/WhatsApp (contatável)", 10),
        ("Porte", "Grande empresa", 15),
        ("Porte", "Média empresa", 12),
        ("Porte", "Pequena empresa", 10),
        ("Porte", "Micro empresa", 5),
        ("Porte", "MEI / Autônomo", 2),
    ]
    for r, (dim, crit, pts) in enumerate(score_table, 4):
        ws2.cell(r, 1, dim).font = Font(bold=True, size=10)
        ws2.cell(r, 2, crit)
        ws2.cell(r, 3, pts).alignment = Alignment(horizontal="center")
        for col in range(1, 4):
            ws2.cell(r, col).border = border

    ws2["A22"] = "Temperatura"
    ws2["B22"] = "Score"
    ws2["A22"].font = ws2["B22"].font = Font(bold=True, color="FFFFFF")
    ws2["A22"].fill = ws2["B22"].fill = HEADER_FILL

    ws2["A23"] = "🔥 Quente"; ws2["B23"] = "≥ 70 pontos"
    ws2["A24"] = "🟡 Morno";  ws2["B24"] = "45 – 69 pontos"
    ws2["A25"] = "❄️ Frio";   ws2["B25"] = "< 45 pontos"
    ws2["A23"].fill = QUENTE_FILL
    ws2["A24"].fill = MORNO_FILL
    ws2["A25"].fill = FRIO_FILL

    ws2.column_dimensions["A"].width = 22
    ws2.column_dimensions["B"].width = 40
    ws2.column_dimensions["C"].width = 10

    # Aba resumo busca
    ws3 = wb.create_sheet("Resumo")
    ws3["A1"] = "Resumo da Coleta"
    ws3["A1"].font = Font(bold=True, size=14, color="1C4587")
    items = [
        ("Prospectador", prospectador),
        ("Segmento buscado", segmento),
        ("Cidade", f"{cidade} – {estado}"),
        ("Data", datetime.now().strftime("%d/%m/%Y %H:%M")),
        ("Total de leads", len(leads)),
        ("🔥 Quentes", sum(1 for l in leads if "Quente" in l.get("classificacao",""))),
        ("🟡 Mornos",  sum(1 for l in leads if "Morno"  in l.get("classificacao",""))),
        ("❄️ Frios",   sum(1 for l in leads if "Frio"   in l.get("classificacao",""))),
        ("Com WhatsApp", sum(1 for l in leads if l.get("is_whatsapp"))),
        ("Com site", sum(1 for l in leads if l.get("site"))),
    ]
    for r, (label, val) in enumerate(items, 3):
        ws3.cell(r, 1, label).font = Font(bold=True)
        ws3.cell(r, 2, val)
        ws3.cell(r, 1).border = border
        ws3.cell(r, 2).border = border
    ws3.column_dimensions["A"].width = 22
    ws3.column_dimensions["B"].width = 30

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf.read()


@app.post("/scrape")
async def start_scrape(req: ScrapeRequest):
    try:
        logger.info(
            "Frontend requested scrape segmento=%s cidade=%s estado=%s max_results=%s",
            req.segmento,
            req.cidade,
            req.estado,
            req.max_results,
        )
        if req.max_results <= 0:
            logger.warning("Rejected scrape request with invalid max_results=%s", req.max_results)
            return JSONResponse({"error": "A quantidade de leads deve ser maior que zero."}, status_code=400)

        job_id = datetime.now().strftime("%Y%m%d%H%M%S%f")
        jobs[job_id] = {"status": "pending", "progress": 0, "total": 0, "leads": [], "log": "Iniciando..."}
        thread = threading.Thread(
            target=scrape_worker,
            args=(job_id, req.segmento, req.cidade, req.estado, req.max_results, req.prospectador),
            daemon=True,
        )
        thread.start()
        logger.info("Scraper thread started job_id=%s", job_id)
        return {"job_id": job_id}
    except Exception:
        logger.exception("Failed to start scrape job")
        traceback.print_exc()
        return JSONResponse({"error": "Falha ao iniciar extração. Veja backend.log para detalhes."}, status_code=500)
@app.get("/status/{job_id}")
async def get_status(job_id: str):
    try:
        job = jobs.get(job_id)
        if job is None:
            logger.warning("Status requested for unknown job_id=%s", job_id)
            return JSONResponse({"error": "Job não encontrado"}, status_code=404)
        return {
            "status": job.get("status"),
            "progress": job.get("progress", 0),
            "total": job.get("total", 0),
            "log": job.get("log", ""),
            "leads_count": len(job.get("leads", [])),
            "leads": job.get("leads", []),
        }
    except Exception:
        logger.exception("Failed to get status for job %s", job_id)
        traceback.print_exc()
        return JSONResponse({"error": "Falha ao consultar status."}, status_code=500)
@app.get("/download/{job_id}")
async def download_xlsx(job_id: str, segmento: str = "", cidade: str = "",
                        estado: str = "", prospectador: str = ""):
    try:
        logger.info("Download requested for job_id=%s", job_id)
        job = jobs.get(job_id)
        if not job or job.get("status") != "done":
            logger.warning("Download rejected for unfinished or missing job_id=%s", job_id)
            return JSONResponse({"error": "Job não concluído"}, status_code=400)
        xlsx_bytes = generate_xlsx(job["leads"], segmento, cidade, estado, prospectador)
        filename = f"leads_{segmento.replace(' ','_')}_{cidade}_{datetime.now().strftime('%Y%m%d')}.xlsx"
        return StreamingResponse(
            io.BytesIO(xlsx_bytes),
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={"Content-Disposition": f'attachment; filename="{filename}"'},
        )
    except Exception:
        logger.exception("Failed to generate XLSX for job %s", job_id)
        traceback.print_exc()
        return JSONResponse({"error": "Falha ao gerar planilha."}, status_code=500)
@app.get("/", response_class=HTMLResponse)
async def home():
    try:
        index_path = Path(__file__).parent / "index.html"
        if index_path.exists():
            return index_path.read_text(encoding="utf-8")
        return "<html><body><h1>Lead Scraper API</h1><p>Backend ativo.</p></body></html>"
    except Exception:
        logger.exception("Failed to serve backend home endpoint")
        traceback.print_exc()
        return JSONResponse({"error": "Falha ao carregar página inicial do backend"}, status_code=500)


@app.get("/health")
async def health():
    return {"ok": True, "db_ready": engine is not None}

@app.get("/history")
async def get_history():
    try:
        logger.info("History requested by frontend")
        ensure_db_ready()
        with db_lock:
            with Session(engine) as session:
                records = session.exec(
                    select(SearchHistory).order_by(SearchHistory.created_at.desc()).limit(50)
                ).all()
                return [r.model_dump() for r in records]
    except Exception:
        logger.exception("Failed to load search history")
        traceback.print_exc()
        return JSONResponse({"error": "Falha ao carregar histórico."}, status_code=500)

multiprocessing.freeze_support()

import uvicorn

logger.info("Starting FastAPI server on 127.0.0.1:8000")
uvicorn.run(
    app,
    host="127.0.0.1",
    port=8000
)
