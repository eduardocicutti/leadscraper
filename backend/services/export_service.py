import io
from datetime import datetime

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.worksheet.datavalidation import DataValidation

from backend.domain.whatsapp import format_phone_br


def generate_xlsx(
    leads: list,
    segmento: str,
    cidade: str,
    estado: str,
    prospectador: str,
) -> bytes:
    wb = Workbook()
    ws = wb.active
    ws.title = "Página1"

    header_fill = PatternFill("solid", start_color="1C4587")
    white_font = Font(bold=True, color="FFFFFF", name="Calibri", size=11)
    thin = Side(style="thin", color="CCCCCC")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    headers = [
        "Responsável",
        "Estágio atual",
        "Nome da empresa",
        "Telefone/WhatsApp",
        "Responsável",
        "Onde encontrou?",
        "Resp",
        "F1",
        "F2",
        "F3",
        "Ramo de atividade",
        "Porte",
        "Nota Google",
        "Avaliações",
        "Score",
        "Temperatura",
        "Link WhatsApp",
        "Endereço",
        "Site",
        "Google Maps",
        "Cidade",
    ]

    col_widths = {
        "A": 15.13,
        "B": 17.0,
        "C": 28.75,
        "D": 17.0,
        "E": 17.63,
        "F": 17.0,
        "G": 7.63,
        "H": 7.0,
        "I": 7.0,
        "J": 7.0,
        "K": 23.75,
        "L": 18.0,
        "M": 12.0,
        "N": 12.0,
        "O": 13.0,
        "P": 13.0,
        "Q": 35.0,
        "R": 40.0,
        "S": 30.0,
        "T": 40.0,
        "U": 15.0,
    }
    for col_letter, width in col_widths.items():
        ws.column_dimensions[col_letter].width = width

    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = white_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = border
    ws.row_dimensions[1].height = 20

    dv_responsavel = DataValidation(
        type="list",
        formula1='"Eduardo,Murilo,Sofia,Ydian,Responsável,Gabriel"',
        showDropDown=False,
    )
    dv_estagio = DataValidation(
        type="list",
        formula1='"Qualificação,Mensagens,Briefing,Perdido"',
        showDropDown=False,
    )
    dv_origem = DataValidation(
        type="list",
        formula1='"Google maps,Indicação,Instagram,Linkedin,Outros,B2B"',
        showDropDown=False,
    )

    ws.add_data_validation(dv_responsavel)
    ws.add_data_validation(dv_estagio)
    ws.add_data_validation(dv_origem)

    quente_fill = PatternFill("solid", start_color="FFE0E0")
    morno_fill = PatternFill("solid", start_color="FFF9C4")
    frio_fill = PatternFill("solid", start_color="E3F2FD")

    max_row = 1002
    for i, lead in enumerate(leads):
        row = i + 2
        clf = lead.get("classificacao", "")
        if "Quente" in clf:
            row_fill = quente_fill
        elif "Morno" in clf:
            row_fill = morno_fill
        else:
            row_fill = frio_fill
        tel_display = format_phone_br(lead.get("telefone"))

        row_values = [
            prospectador,
            "Qualificação",
            lead.get("nome", ""),
            tel_display,
            "",
            "Google maps",
            False,
            False,
            False,
            False,
            lead.get("categoria", ""),
            lead.get("porte", ""),
            lead.get("nota"),
            lead.get("avaliacoes"),
            lead.get("score", 0),
            clf,
            lead.get("whatsapp_link", ""),
            lead.get("endereco", ""),
            lead.get("site", ""),
            lead.get("url_maps", ""),
            f"{lead.get('cidade', '')}/{lead.get('estado', '')}",
        ]

        for col, val in enumerate(row_values, 1):
            cell = ws.cell(row=row, column=col, value=val)
            cell.fill = row_fill
            cell.border = border
            cell.alignment = Alignment(vertical="center", wrap_text=False)
            if col in (7, 8, 9, 10):
                cell.alignment = Alignment(horizontal="center", vertical="center")
            if col == 3:
                cell.font = Font(bold=True, name="Calibri", size=10)
            else:
                cell.font = Font(name="Calibri", size=10)

        ws.row_dimensions[row].height = 18

        dv_responsavel.sqref = f"A2:A{max_row}"
        dv_estagio.sqref = f"B2:B{max_row}"
        dv_origem.sqref = f"F2:F{max_row}"

    ws.freeze_panes = "A2"

    ws2 = wb.create_sheet("Score")
    ws2["A1"] = "SISTEMA DE SCORE"
    ws2["A1"].font = Font(bold=True, size=14, color="1C4587")
    ws2.merge_cells("A1:C1")

    ws2["A3"] = "Dimensão"
    ws2["B3"] = "Critério"
    ws2["C3"] = "Pontos"
    for col in range(1, 4):
        ws2.cell(3, col).font = Font(bold=True, color="FFFFFF")
        ws2.cell(3, col).fill = header_fill
        ws2.cell(3, col).alignment = Alignment(horizontal="center")

    score_table = [
        ("Presença Digital", "Sem site (oportunidade clara de presença digital)", 25),
        ("Presença Digital", "Com site (pode querer sistema/app/redesign)", 5),
        ("Avaliações Google", "≥ 100 avaliações", 20),
        ("Avaliações Google", "50 – 99 avaliações", 15),
        ("Avaliações Google", "10 – 49 avaliações", 10),
        ("Avaliações Google", "1 – 9 avaliações", 5),
        ("Nota Google", "≥ 4.5 ⭐ (empresa de qualidade)", 15),
        ("Nota Google", "4.0 – 4.4 ⭐", 10),
        ("Nota Google", "3.5 – 3.9 ⭐", 5),
        ("Fit de Segmento", "Segmento com alto potencial digital", 20),
        ("Fit de Segmento", "Tem telefone/WhatsApp (contatável)", 10),
        ("Porte", "Grande empresa", 15),
        ("Porte", "Média empresa", 12),
        ("Porte", "Pequena empresa", 10),
        ("Porte", "Micro empresa", 5),
        ("Porte", "MEI / Autônomo", 2),
    ]
    for r, (dim, crit, pts) in enumerate(score_table, 4):
        ws2.cell(r, 1, dim).font = Font(bold=True, size=10)
        ws2.cell(r, 2, crit)
        ws2.cell(r, 3, pts).alignment = Alignment(horizontal="center")
        for col in range(1, 4):
            ws2.cell(r, col).border = border

    ws2["A22"] = "Temperatura"
    ws2["B22"] = "Score"
    ws2["A22"].font = ws2["B22"].font = Font(bold=True, color="FFFFFF")
    ws2["A22"].fill = ws2["B22"].fill = header_fill

    ws2["A23"] = "🔥 Quente"
    ws2["B23"] = "≥ 70 pontos"
    ws2["A24"] = "🟡 Morno"
    ws2["B24"] = "45 – 69 pontos"
    ws2["A25"] = "❄️ Frio"
    ws2["B25"] = "< 45 pontos"
    ws2["A23"].fill = quente_fill
    ws2["A24"].fill = morno_fill
    ws2["A25"].fill = frio_fill

    ws2.column_dimensions["A"].width = 22
    ws2.column_dimensions["B"].width = 40
    ws2.column_dimensions["C"].width = 10

    ws3 = wb.create_sheet("Resumo")
    ws3["A1"] = "Resumo da Coleta"
    ws3["A1"].font = Font(bold=True, size=14, color="1C4587")
    items = [
        ("Prospectador", prospectador),
        ("Segmento buscado", segmento),
        ("Cidade", f"{cidade} – {estado}"),
        ("Data", datetime.now().strftime("%d/%m/%Y %H:%M")),
        ("Total de leads", len(leads)),
        ("🔥 Quentes", sum(1 for lead in leads if "Quente" in lead.get("classificacao", ""))),
        ("🟡 Mornos", sum(1 for lead in leads if "Morno" in lead.get("classificacao", ""))),
        ("❄️ Frios", sum(1 for lead in leads if "Frio" in lead.get("classificacao", ""))),
        ("Com WhatsApp", sum(1 for lead in leads if lead.get("is_whatsapp"))),
        ("Com site", sum(1 for lead in leads if lead.get("site"))),
    ]
    for r, (label, val) in enumerate(items, 3):
        ws3.cell(r, 1, label).font = Font(bold=True)
        ws3.cell(r, 2, val)
        ws3.cell(r, 1).border = border
        ws3.cell(r, 2).border = border
    ws3.column_dimensions["A"].width = 22
    ws3.column_dimensions["B"].width = 30

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf.read()
